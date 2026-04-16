import io
import os
import re
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pdfplumber
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DEFAULT_TEMPLATE_PATH = DATA_DIR / "Bulk_Load_Template.xlsx"
DEFAULT_LOOKUP_PATH = DATA_DIR / "Address_Lookup.xlsx"
DEFAULT_LOCATION_ID = "25613725"
DATE_FMT = "%m-%d-%Y"
TIME_FMT = "%H:%M"
ADDRESS_LOOKUP = None


def load_address_lookup(address_lookup_path: str) -> pd.DataFrame:
    df = pd.read_excel(address_lookup_path, sheet_name="Sheet1")
    df["Street_Only"] = (
        df["Address"]
        .astype(str)
        .str.upper()
        .str.replace(".", "", regex=False)
        .str.replace(",", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.replace(" ROAD", " RD", regex=False)
        .str.strip()
    )
    return df


def lookup_address_id(street_address: str) -> str:
    global ADDRESS_LOOKUP

    if ADDRESS_LOOKUP is None:
        raise RuntimeError("Address lookup table not loaded.")

    if not street_address or street_address.strip() == "" or street_address == "N/A":
        return "N/A"

    cleaned = re.sub(
        r"\b(UNIT|APT|SUITE|STE|FLOOR|FL)\s*#?\s*\d+\b",
        "",
        street_address,
        flags=re.IGNORECASE,
    )

    pdf_addr = (
        cleaned.upper()
        .replace(".", "")
        .replace(",", " ")
        .replace(" ROAD", " RD")
        .strip()
    )
    pdf_addr = re.sub(r"\s+", " ", pdf_addr)

    parts = pdf_addr.split()
    if len(parts) < 2:
        return "N/A"

    pdf_number = parts[0]
    pdf_street = parts[1]

    mask = (
        ADDRESS_LOOKUP["Street_Only"].str.contains(pdf_number, na=False)
        & ADDRESS_LOOKUP["Street_Only"].str.contains(pdf_street, na=False)
    )

    matches = ADDRESS_LOOKUP[mask]
    return str(matches.iloc[0]["ID"]) if not matches.empty else "N/A"


def extract_text_from_pdf(pdf_path: str) -> str:
    full_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
    return full_text


def parse_header(full_text: str) -> dict:
    load_id = re.search(
        r"(?:Num[eé]ro\s+de\s+route|Load\s+ID)[:\s]+([\d]+)",
        full_text,
        re.IGNORECASE,
    )
    distance = re.search(
        r"Distance\s*\(?(km|mi)?\)?[:\s]+([\d,]+\.?\d*)",
        full_text,
        re.IGNORECASE,
    )
    return {
        "Load ID": load_id.group(1).strip() if load_id else "N/A",
        "Distance": distance.group(2).strip() if distance else "N/A",
    }


def parse_dt(date_str: str, time_str: str) -> datetime:
    return datetime.strptime(f"{date_str} {time_str}", f"{DATE_FMT} {TIME_FMT}")


def fmt_date(dt: datetime) -> str:
    return dt.strftime(DATE_FMT)


def fmt_time(dt: datetime) -> str:
    return dt.strftime(TIME_FMT)


def normalize_pdf_date(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%d-%b-%Y")
        return dt.strftime("%m-%d-%Y")
    except Exception:
        return date_str


def parse_stops(full_text: str) -> list:
    stops = []
    stop_blocks = re.split(r"(Arrêt\s*/\s*Stop\s+\d+[:\s]*)", full_text, flags=re.IGNORECASE)

    for i in range(1, len(stop_blocks), 2):
        header = stop_blocks[i]
        block = stop_blocks[i + 1]

        stop_num_match = re.search(r"\d+", header)
        if not stop_num_match:
            continue
        stop_num = stop_num_match.group(0)

        stop_type = re.search(
            r"(?:DC|Store|Warehouse|Distribution Center|Plant)",
            block,
            re.IGNORECASE,
        )
        arrival = re.search(r"Arri(?:vée|val).*?([\d]{1,2}-\w{3}-\d{4}\s+[\d:]+)", block)
        departure = re.search(r"D[ée]part.*?([\d]{1,2}-\w{3}-\d{4}\s+[\d:]+)", block)
        postal_city = re.search(r"([A-Z]\d[A-Z]\d[A-Z]\d),?\s+([A-Z][A-Z\s]+)", block)
        us_address = re.search(r"([A-Z][a-zA-Z\s]+),\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)", block)
        street = re.search(
            r"(?m)^(?:\d{1,6})\s+[A-Z0-9 .'-]+(?:MALL|ROAD|RD|STREET|ST|AVENUE|AVE|BLVD|BOULEVARD|DRIVE|DR|WAY|LANE|LN|COURT|CT|HWY|HIGHWAY)\b.*$",
            block,
        )

        store_name = ""
        if street:
            before_street = block[: block.find(street.group(0))]
            lines = [line.strip() for line in before_street.split("\n") if line.strip()]
            if lines:
                last_line = lines[-1]
                if not re.match(r"^\d{1,2}-\w{3}-\d{4}|Arri|Départ", last_line):
                    store_name = last_line

        supplier = re.search(r"(?:Fournisseur|Customer).*?(.+)", block)
        po = re.search(r"(?:Purchase Order|Bon de commande)\s*([\d\s]+)", block)
        shipment = re.search(r"(\d+)\s+(\d+)\s+([\d.]+)\s+([\d.]+)", block)

        street_address = street.group(0).strip() if street else "N/A"
        address_id = lookup_address_id(street_address)

        def split_dt(val):
            if not val:
                return "N/A", "N/A"
            parts = val.split()
            return parts[0], parts[1]

        arr_date, arr_time = split_dt(arrival.group(1) if arrival else None)
        dep_date, dep_time = split_dt(departure.group(1) if departure else None)

        arr_date = normalize_pdf_date(arr_date)
        dep_date = normalize_pdf_date(dep_date)

        stops.append(
            {
                "Stop #": stop_num,
                "Stop Type": stop_type.group(0) if stop_type else "N/A",
                "Store / Facility": store_name or "N/A",
                "Supplier": supplier.group(1).strip() if supplier else "N/A",
                "Street Address": street_address,
                "Address ID": address_id,
                "City": postal_city.group(2).strip()
                if postal_city
                else (us_address.group(1).strip() if us_address else "N/A"),
                "Postal Code": postal_city.group(1).strip()
                if postal_city
                else (us_address.group(3).strip() if us_address else "N/A"),
                "Arrival Date": arr_date,
                "Arrival Time": arr_time,
                "Departure Date": dep_date,
                "Departure Time": dep_time,
                "Purchase Order": po.group(1).strip() if po else "N/A",
                "Pallets": shipment.group(1) if shipment else "N/A",
                "Pieces": shipment.group(2) if shipment else "N/A",
                "Weight": shipment.group(4) if shipment else "N/A",
            }
        )

    return stops


def clear_existing_rows(template_path: str) -> None:
    wb = load_workbook(template_path)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError(f"'Sheet1' not found in {template_path}")

    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    wb.save(template_path)


def save_rows_to_bulk_template(rows: list, template_path: str) -> None:
    wb = load_workbook(template_path)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError(f"'Sheet1' not found in {template_path}")

    ws = wb["Sheet1"]
    sheet_columns = {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value
    }

    first_empty_row = None
    for r in range(2, ws.max_row + 200):
        row_is_empty = True
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, "", " "):
                row_is_empty = False
                break
        if row_is_empty:
            first_empty_row = r
            break

    if first_empty_row is None:
        first_empty_row = ws.max_row + 1

    for offset, row_data in enumerate(rows):
        target_row = first_empty_row + offset
        for col_name, col_idx in sheet_columns.items():
            ws.cell(row=target_row, column=col_idx, value=row_data.get(col_name))

    wb.save(template_path)


def build_excel_rows(header: dict, stops: list) -> list:
    rows = []
    load_id = header["Load ID"]
    total_stops = len(stops)

    first_stop = stops[0]
    last_stop = stops[-1]

    first_arrival_dt = parse_dt(first_stop["Arrival Date"], first_stop["Arrival Time"])
    start_dt = first_arrival_dt - timedelta(hours=1)

    rows.append(
        {
            "LoadId": load_id,
            "Scheduled Start Date": fmt_date(start_dt),
            "Schedule Start Time": fmt_time(start_dt),
            "Scheduled Finish Date": fmt_date(start_dt),
            "Schedule Finish Time": fmt_time(start_dt),
            "Schedule start Location Id": DEFAULT_LOCATION_ID,
            "Schedule end Location Id": DEFAULT_LOCATION_ID,
            "Stop Type (*)": "START",
            "Stop Sequence": "0",
            "Scheduled Date": fmt_date(start_dt),
            "Scheduled Time": fmt_time(start_dt),
            "Planned No of Stops": total_stops,
            "Planned Total KM": header["Distance"],
            "weight": "",
            "Pieces": "",
            "Load Type e.g. Box/Pallet": "",
        }
    )

    for idx, stop in enumerate(stops):
        next_stop = stops[idx + 1] if idx + 1 < len(stops) else None
        rows.append(
            {
                "LoadId": load_id,
                "Scheduled Start Date": stop["Arrival Date"],
                "Schedule Start Time": stop["Arrival Time"],
                "Scheduled Finish Date": stop["Departure Date"],
                "Schedule Finish Time": stop["Departure Time"],
                "Schedule start Location Id": stop["Address ID"],
                "Schedule end Location Id": next_stop["Address ID"] if next_stop else DEFAULT_LOCATION_ID,
                "Stop Type (*)": stop["Stop Type"],
                "Stop Sequence": stop["Stop #"],
                "Scheduled Date": stop["Arrival Date"],
                "Scheduled Time": stop["Arrival Time"],
                "Planned No of Stops": total_stops,
                "Planned Total KM": header["Distance"],
                "weight": stop["Weight"],
                "Pieces": stop["Pieces"],
                "Load Type e.g. Box/Pallet": stop["Pallets"],
            }
        )

    last_dep_dt = parse_dt(last_stop["Departure Date"], last_stop["Departure Time"])
    end_dt = last_dep_dt + timedelta(hours=1)

    rows.append(
        {
            "LoadId": load_id,
            "Scheduled Start Date": fmt_date(last_dep_dt),
            "Schedule Start Time": fmt_time(last_dep_dt),
            "Scheduled Finish Date": fmt_date(last_dep_dt),
            "Schedule Finish Time": fmt_time(end_dt),
            "Schedule start Location Id": DEFAULT_LOCATION_ID,
            "Schedule end Location Id": DEFAULT_LOCATION_ID,
            "Stop Type (*)": "END",
            "Stop Sequence": str(total_stops + 1),
            "Scheduled Date": fmt_date(last_dep_dt),
            "Scheduled Time": fmt_time(last_dep_dt),
            "Planned No of Stops": total_stops,
            "Planned Total KM": header["Distance"],
            "weight": "",
            "Pieces": "",
            "Load Type e.g. Box/Pallet": "",
        }
    )

    return rows


def process_load_tender(pdf_path: str, template_path: str) -> None:
    text = extract_text_from_pdf(pdf_path)
    header = parse_header(text)
    stops = parse_stops(text)
    if not stops:
        return
    rows = build_excel_rows(header, stops)
    save_rows_to_bulk_template(rows, template_path)


@app.get("/")
def home():
    return jsonify({"ok": True, "message": "Tender processor backend is running."})


@app.get("/health")
def health():
    return jsonify({"status": "healthy"})


@app.post("/process-tenders")
def process_tenders():
    global ADDRESS_LOOKUP

    pdf_files = request.files.getlist("pdf_files")
    if not pdf_files:
        return jsonify({"error": "No PDF files uploaded."}), 400

    temp_dir = tempfile.mkdtemp(prefix="tenders_")
    try:
        working_template = Path(temp_dir) / "Bulk_Load_Output.xlsx"
        shutil.copy(DEFAULT_TEMPLATE_PATH, working_template)

        ADDRESS_LOOKUP = load_address_lookup(str(DEFAULT_LOOKUP_PATH))
        clear_existing_rows(str(working_template))

        saved_pdf_paths = []
        for pdf in pdf_files:
            if not pdf.filename:
                continue
            pdf_path = Path(temp_dir) / pdf.filename
            pdf.save(pdf_path)
            saved_pdf_paths.append(str(pdf_path))

        if not saved_pdf_paths:
            return jsonify({"error": "Uploaded PDF list was empty."}), 400

        saved_pdf_paths.sort()
        for pdf_path in saved_pdf_paths:
            process_load_tender(pdf_path, str(working_template))

        with open(working_template, "rb") as f:
            output_bytes = io.BytesIO(f.read())
        output_bytes.seek(0)

        return send_file(
            output_bytes,
            as_attachment=True,
            download_name="Bulk_Load_Output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
